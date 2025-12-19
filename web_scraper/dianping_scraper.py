import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from utils.login_utils.browser_setup import create_driver
from utils.login_utils.cookies_manager import get_or_load_cookies
from utils.file_utils.file_path_and_creat_folder import create_output_folder
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import csv
import random
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from utils.file_utils.load_datas import resource_path, load_json_file



class Dpscraper:
    def __init__(self, key_word, page_count, city, log_signal=None, thread_instance=None):
        self.key_word = key_word.strip()  # 去除关键词前后空格
        self.page_count = int(page_count)
        self.log_signal = log_signal
        self.city = city  # 添加城市参数
        self.thread_instance = thread_instance
        self.driver = create_driver(log_signal=self.log_signal, headless=False)

        # 调用方法生成文件名
        self.filename = self.generate_filename()

        # 尝试获取或加载 cookies
        login_url = "https://www.dianping.com/chengdu"  # 替换为实际的登录 URL
        get_or_load_cookies(self.driver, login_url, "dianping", self.log_signal, login_click_xpath="//*[@id='top-nav']/div/div[2]/span[1]/a")


    def generate_filename(self):
        """根据参数生成唯一的文件名"""
        # 用关键词、页数和地区生成文件名
        base_name = f"{self.city}_{self.key_word}_{self.page_count}页"

        # 确保文件名合法，不包含特殊字符
        filename = base_name.replace(" ", "_").replace("/", "_")
        return filename


    def log_message(self, message):
        """日志输出函数，如果 log_signal 存在，则发送信号；否则直接打印"""
        if self.log_signal:
            self.log_signal.emit(message)
        else:
            print(message)


    def _is_not_found(self):
        """检查页面是否显示未找到内容的标志"""
        try:
            self.driver.find_element(By.CLASS_NAME, "not-found-right")
            return True
        except NoSuchElementException:
            return False


    def scrape(self, custom_base_dir=None):
        """开始爬取"""

        # Step 1: 从 JSON 文件加载城市链接映射
        try:
            city_links = load_json_file(resource_path("utils/dianping_city_link.json"))
        except FileNotFoundError as e:
            raise FileNotFoundError(f"无法加载城市链接映射文件: {str(e)}")

        # 验证城市名称
        city_name = self.city.strip() if self.city else None
        if not city_name:
            raise ValueError("城市名称不能为空，请输入有效的城市名称！")

        if city_name not in city_links:
            supported_cities = ", ".join(city_links.keys())  # 列出支持的城市
            raise ValueError(
                f"未找到城市 {city_name} 的链接，请检查输入！\n"
                f"支持的城市有：{supported_cities}"
            )

        # 获取城市对应的 URL
        city_url = city_links[city_name]

        # 打开城市页面
        self.driver.get(city_url)
        self.log_message(f"成功打开城市页面: {city_url}")
        time.sleep(3)  # 等待页面加载

        # Step 2: 显式等待并输入关键词搜索
        try:
            # 等待搜索输入框可见
            search_box = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='J-search-input']"))
            )
            search_box.clear()  # 清空搜索框
            search_box.send_keys(self.key_word)
            self.log_message(f"已输入关键词: {self.key_word}")

            # 等待搜索按钮可点击
            search_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='J-all-btn']"))
            )
            search_button.click()
            self.log_message("已点击搜索按钮，等待新标签页加载")

            # 显式等待新标签页打开
            WebDriverWait(self.driver, 10).until(
                lambda driver: len(driver.window_handles) > 1
            )
            self.log_message("新标签页已打开")

            # 切换到新标签页
            new_tab = self.driver.window_handles[-1]  # 获取最新的标签页句柄
            self.driver.switch_to.window(new_tab)
            self.log_message(f"已切换到新标签页: {self.driver.current_url}")

        except TimeoutException:
            raise TimeoutException("未能定位到搜索框或搜索按钮，请检查页面结构")

        # Step 3: 等待搜索结果页面加载完成并提取分页 URL 模板
        # 获取当前标签页的 URL
        search_url = self.driver.current_url
        self.log_message(f"当前搜索结果的 URL: {search_url}")

        if "/p" in search_url:
            base_url = search_url.split("/p")[0] + "/p{}"
        else:
            base_url = search_url + "/p{}"
        self.log_message(f"分页 URL 模板: {base_url}")

        # 创建输出文件夹
        output_folder = create_output_folder(base_url, custom_base_dir=custom_base_dir)
        self.log_message(f"输出文件夹创建成功: {output_folder}")

        all_links = []  # 用于存储店铺链接
        results = []  # 用于存储爬取结果

        # Step 4: 收集所有链接
        try:
            pages = list(range(1, int(self.page_count) + 1))
        except ValueError:
            raise ValueError(f"page_count 必须是整数类型，当前值为: {self.page_count}")
        self.log_message(f"page_count 值: {self.page_count}, 类型: {type(self.page_count)}")

        if len(pages) > 1:
            random.shuffle(pages)  # 只有多页时打乱页面顺序

        for page in pages:
            url = base_url.format(page)
            self.log_message(f"正在访问: {url}")
            self.driver.get(url)
            time.sleep(random.uniform(4.5, 5.5))  # 随机延时
            random_scroll(self.driver)

            try:
                # 检查是否存在 "not-found-right" 标志
                if self._is_not_found():
                    self.log_message("未找到更多内容，跳过该页面。")
                    continue  # 跳过无效页面，但继续处理剩余页面

                # 等待页面加载完成
                WebDriverWait(self.driver, 6).until(
                    EC.presence_of_all_elements_located(
                        (By.XPATH, "//li[contains(@class, '') and .//div[@class='txt']]"))
                )

                # 查找所有店铺的列表
                shops = self.driver.find_elements(By.XPATH, "//li[contains(@class, '') and .//div[@class='txt']]")
                for shop in shops:
                    try:
                        # 提取店铺链接
                        link = wait_for_element(shop, By.XPATH, ".//div[@class='tit']/a").get_attribute("href")
                        # 提取类别和位置
                        tags = shop.find_elements(By.XPATH, ".//div[@class='tag-addr']//span[@class='tag']")
                        category = tags[0].text if len(tags) > 0 else None
                        location = tags[1].text if len(tags) > 1 else None
                        # self.log_message(f"{category}\n{location}")

                        # 将信息存储到链接字典
                        if link:
                            all_links.append({"链接": link, "类别": category, "所在区域": location})
                    except Exception as e:
                        self.log_message(f"提取店铺链接或基本信息时发生错误: {e}")

            except TimeoutException:
                self.log_message(f"页面加载超时: {url}")
            except Exception as e:
                self.log_message(f"爬取时发生错误: {e}")

        # Step 5: 转换链接为旧版链接
        def convert_to_old_link(link):
            """将新版店铺链接转换为旧版链接"""
            if "shop/" in link:
                shop_id = link.split("shop/")[-1]
                return f"https://www.dianping.com/shopold/pc?shopuuid={shop_id}"
            return link

        # Step 6: 逐个访问旧版链接并提取内容
        # 打乱链接顺序
        if not all_links:
            self.log_message("链接列表为空，跳过爬取流程。")
            return  # 或者 raise ValueError("链接列表为空")

        random.shuffle(all_links)

        for link_item in all_links:
            # 转换链接
            old_link = convert_to_old_link(link_item["链接"])

            if old_link:  # 检查转换后的链接是否有效
                self.log_message(f"正在访问店铺: {old_link}")
                self.driver.get(old_link)  # 传入字符串链接
                time.sleep(random.uniform(1.5, 2.5))  # 随机延时
                random_scroll(self.driver)

                try:
                    # 初始化店铺信息字段，设置默认值
                    name = review_count = avg_price = address = "N/A"
                    scores = {}
                    branch_info = None  # 初始化分店信息

                    # 使用通用等待函数提取店铺信息
                    try:
                        name_element = wait_for_element(self.driver, By.XPATH, "//h1[@class='shop-name']")
                        if name_element:
                            name = name_element.text.strip()  # 使用 .text 获取文本
                            self.log_message(f"店铺名称: {name}")

                            # 去除 "手机扫码 优惠买单" 部分
                            name = name.replace("手机扫码 优惠买单", "").strip()

                            # 如果店铺名称中包含"其它X家分店"，提取分店信息
                            match = re.search(r"其它(\d+)家分店", name)
                            if match:
                                branch_info = match.group(1)  # 提取数字部分
                                self.log_message(f"分店数量: {branch_info}")
                                # 去除分店信息
                                name = re.sub(r"其它\d+家分店", "", name).strip()  # 删除整个分店信息文本
                            else:
                                branch_info = "0"  # 默认值为 0（无分店信息）

                            self.log_message(f"最终店铺名称: {name}")
                        else:
                            self.log_message("未找到店铺名称元素")
                    except Exception as e:
                        self.log_message(f"提取店铺名称时发生错误: {e}")

                    try:
                        review_count = wait_for_element(self.driver, By.XPATH, "//span[@id='reviewCount']").text.strip()
                        # 使用正则提取数字部分
                        review_count = re.search(r"\d+", review_count.replace(",", "")).group()
                        self.log_message(f"评价数: {review_count}")
                    except Exception as e:
                        self.log_message(f"未找到评价数元素: {e}")

                    try:
                        avg_price = wait_for_element(self.driver, By.XPATH, "//span[@id='avgPriceTitle']").text.strip()
                        # 使用正则提取数字部分
                        avg_price = re.search(r"\d+", avg_price).group()
                        self.log_message(f"人均消费: {avg_price}")
                    except Exception as e:
                        self.log_message(f"未找到人均消费元素: {e}")

                    try:
                        address = wait_for_element(
                            self.driver, By.XPATH, "//div[@class='expand-info address']//span[@id='address']"
                        ).text.strip()
                        self.log_message(f"地址: {address}")
                    except Exception as e:
                        self.log_message(f"未找到地址元素: {e}")

                    # 动态提取评分项
                    try:
                        comment_score_element = wait_for_element(self.driver, By.ID, "comment_score")
                        score_items = comment_score_element.find_elements(By.CLASS_NAME, "item")

                        # 动态提取所有评分项
                        for item in score_items:
                            text = item.text.strip()  # 例如 "口味: 8.9" 或 "交通: 7.5"
                            if ":" in text:
                                key, value = text.split(":", 1)  # 使用冒号分割
                                key = key.strip()  # 去除多余空格
                                value = value.strip()  # 去除多余空格
                                scores[key] = value  # 动态存储评分项
                                self.log_message(f"找到评分项: {key} -> {value}")
                            else:
                                self.log_message(f"评分项格式不符合预期: {text}")
                    except Exception as e:
                        self.log_message(f"未找到评分元素或提取失败: {e}")

                    # 存储信息
                    result = OrderedDict()
                    result["名称"] = name
                    result["类别"] = link_item.get("类别", "N/A")
                    result["所在区域"] = link_item.get("所在区域", "N/A")
                    result["评价数"] = review_count
                    result["人均消费"] = avg_price
                    result["详细地址"] = address
                    result["区域分店数量"] = branch_info if branch_info else "0"

                    # 动态添加评分项到结果
                    if scores:
                        result.update(scores)  # 动态添加评分项

                    # 最后添加链接
                    result["链接"] = old_link

                    results.append(result)
                    self.log_message(f"已成功提取店铺信息: {result}")

                except StaleElementReferenceException as e:
                    self.log_message(f"发生元素引用错误，跳过该店铺: {old_link} | 错误信息: {e}")
                except Exception as e:
                    self.log_message(f"提取店铺信息时发生错误 ({old_link}): {e}")

        # 保存结果
        self.log_message(f"最终提取到的店铺数量: {len(results)}")
        if len(results) > 0:
            save_to_csv(results, output_folder, self.filename)
            save_to_excel(results, output_folder, self.filename)
        else:
            self.log_message("没有数据保存到文件中")
        self.driver.quit()
        # return results, output_folder


def save_to_csv(data, output_folder, filename):
    """将数据保存为 CSV 文件到指定文件夹"""
    if not data:
        print("数据为空，无法保存")
        return

    # 为文件名添加扩展名，如果没有
    if not filename.endswith('.csv'):
        filename += '.csv'

    # 动态获取所有字段名
    keys = set()
    for item in data:
        keys.update(item.keys())  # 更新字段名集合
    keys = list(keys)  # 转换为列表

    output_path = os.path.join(output_folder, filename)  # 构建完整路径
    with open(output_path, mode="w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=keys)
        writer.writeheader()
        writer.writerows(data)
    print(f"数据已保存为 CSV 文件：{output_path}")


def save_to_excel(data, output_folder, filename):
    """将数据保存为 Excel 文件到指定文件夹并根据内容调整列宽"""
    try:
        # 为文件名添加扩展名，如果没有
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        # 检查输出文件夹是否存在
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # 构建完整路径
        output_path = os.path.join(output_folder, filename)

        # 将数据转换为 pandas DataFrame
        df = pd.DataFrame(data)

        # 将数据保存为 Excel 文件
        df.to_excel(output_path, index=False, engine="openpyxl")

        # 调整列宽
        adjust_excel_column_width(output_path)

        print(f"数据已成功保存为 Excel 文件：{output_path}")
    except Exception as e:
        print(f"保存 Excel 文件时发生错误: {e}")


def adjust_excel_column_width(filepath):
    """根据内容调整 Excel 表格的列宽（考虑中文字符宽度）"""
    try:
        # 加载 Excel 文件
        wb = load_workbook(filepath)
        ws = wb.active  # 获取活动工作表

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)  # 获取列字母
            for cell in col:
                try:
                    if cell.value:
                        # 获取单元格内容并计算宽度
                        cell_value = str(cell.value)
                        # 中文字符宽度按 2，其他字符宽度按 1 计算
                        length = sum(2 if re.match(r"[\u4e00-\u9fff]", char) else 1 for char in cell_value)
                        max_length = max(max_length, length)
                        # 确保数字以数值格式存储
                        if isinstance(cell.value, str):
                            if cell.value.replace(".", "", 1).isdigit():
                                cell.value = float(cell.value) if "." in cell.value else int(cell.value)
                except Exception as e:
                    print(f"无法处理单元格 {cell.coordinate}: {e}")
            # 设置列宽（增加缓冲宽度）
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        # 保存调整后的 Excel 文件
        wb.save(filepath)
        wb.close()
        print("Excel 列宽已根据内容调整完成")
    except Exception as e:
        print(f"调整 Excel 列宽时发生错误: {e}")
        print(f"调整 Excel 列宽时发生错误: {e}")





# 等待函数
def wait_for_element(shop, by, locator, timeout=3):
    return WebDriverWait(shop, timeout).until(
        EC.presence_of_element_located((by, locator))
    )


def random_scroll(driver, total_duration=2, scroll_interval=0.5):
    """
    随机小幅滚动页面，并模拟鼠标活动，增强人类行为模拟。

    参数:
    - driver: Selenium WebDriver 实例
    - total_duration: 滚动的总时长（秒）
    - scroll_interval: 每次滚动之间的间隔时间（秒）
    """
    # 获取浏览器窗口大小和页面高度
    window_size = driver.get_window_size()  # 获取窗口的宽和高
    page_height = driver.execute_script("return document.body.scrollHeight")  # 获取页面总高度
    viewport_height = window_size["height"]  # 浏览器可见区域高度

    start_time = time.time()

    while time.time() - start_time < total_duration:
        try:
            # 随机生成滚动的像素范围
            scroll_amount = random.randint(-300, 300)  # 向上或向下滚动
            driver.execute_script(f"window.scrollBy(0, {scroll_amount});")

            # 当前滚动位置
            scroll_position = driver.execute_script("return window.scrollY")

            # 随机生成鼠标移动坐标 (限制在可见区域范围内)
            random_x = random.randint(0, window_size["width"] - 50)  # 减去一定值避免到窗口边缘
            random_y = random.randint(0, min(viewport_height - 50, page_height - scroll_position - 50))

            # 确保生成的坐标在有效范围内
            if random_y > viewport_height:
                random_y = viewport_height - 50  # 防止鼠标超出窗口范围

            # 模拟鼠标移动到随机位置
            ActionChains(driver).move_by_offset(random_x, random_y).perform()

            # 随机暂停一段时间
            time.sleep(scroll_interval + random.uniform(0, 0.5))

        except Exception as e:
            print(f"滚动或鼠标模拟过程中发生错误: {e}")
            break  # 停止执行，以免重复错误

    print("随机滚动和鼠标模拟完成")


